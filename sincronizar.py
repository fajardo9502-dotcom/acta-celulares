import os
import openpyxl
from openpyxl import load_workbook

# Configura las rutas exactas de tus dos archivos de Excel
RUTA_BASE_FABI = r"C:\Users\1030650138\OneDrive - Colombiana de Comercio S.A\Base Fabi.xlsx"          # El archivo que tú alimentas con los cargos nuevos
RUTA_ACTAS_ENTREGA = r"C:\Users\1030650138\OneDrive - Colombiana de Comercio S.A\base actas de entrega.xlsx"     # El archivo de las actas de entrega

def sincronizar_actas_desde_base_fabi():
    print("🔄 Iniciando sincronización masiva desde Base Fabi...")
    
    if not os.path.exists(RUTA_BASE_FABI) or not os.path.exists(RUTA_ACTAS_ENTREGA):
        print("❌ Error: Uno de los archivos de Excel no fue encontrado. Verifica las rutas.")
        return

    # ==========================================
    # PASO 1: LEER LA BASE FABI Y CREAR EL MAPA
    # ==========================================
    wb_fabi = load_workbook(RUTA_BASE_FABI, data_only=True)
    ws_fabi = wb_fabi.active
    
    # Creamos un diccionario en memoria para guardar las equivalencias rápidamente
    # Estructura: { 'TIPO LOGIA': [Tipo Cargo, Cargo Linea, Tipo] }
    mapa_cargos = {}
    
    # Asumimos que la fila 1 son encabezados y los datos empiezan en la 2
    for row in range(2, ws_fabi.max_row + 1):
        # NOTA: Ajusta el número de las columnas según tu archivo Base Fabi real
        tipo_logia_fabi = ws_fabi.cell(row=row, column=1).value  # Columna A (1) es la llave
        
        # Supongamos estos números de columna de ejemplo para la Base Fabi (Ajústalos si varían):
        tipo_cargo_fabi = ws_fabi.cell(row=row, column=2).value  
        cargo_linea_fabi = ws_fabi.cell(row=row, column=3).value 
        tipo_fabi = ws_fabi.cell(row=row, column=4).value        
        
        if tipo_logia_fabi:
            llave = str(tipo_logia_fabi).strip().upper()
            mapa_cargos[llave] = {
                "tipo_cargo": tipo_cargo_fabi,
                "cargo_linea": cargo_linea_fabi,
                "tipo": tipo_fabi
            }
            
    wb_fabi.close()
    print(f"📖 Base Fabi procesada. Se cargaron {len(mapa_cargos)} cargos en memoria.")

    # ===================================================
    # PASO 2: BUSCAR ACTAS VACÍAS Y ACTUALIZARLAS
    # ===================================================
    wb_actas = load_workbook(RUTA_ACTAS_ENTREGA)
    ws_actas = wb_actas.active
    contador_actualizaciones = 0

    for row in range(2, ws_actas.max_row + 1):
        # Capturamos la llave en el acta: Columna AA es la 27
        tipo_logia_acta = ws_actas.cell(row=row, column=27).value
        
        # Revisamos si las columnas de destino están vacías
        # P = 16, Q = 17, Z = 26
        cargo_linea_actual = ws_actas.cell(row=row, column=16).value
        tipo_actual = ws_actas.cell(row=row, column=17).value
        tipo_cargo_actual = ws_actas.cell(row=row, column=26).value

        # 🎯 CONDICIÓN: Si el acta tiene Tipo Logia (AA) pero le faltan los datos del cruce
        if tipo_logia_acta and (not cargo_linea_actual or not tipo_actual or not tipo_cargo_actual):
            llave_acta = str(tipo_logia_acta).strip().upper()
            
            # Buscamos si esa llave YA existe en la Base Fabi que leímos en el Paso 1
            if llave_acta in mapa_cargos:
                datos_nuevos = mapa_cargos[llave_acta]
                
                # ¡INYECTAMOS LOS DATOS ACTUALIZADOS!
                ws_actas.cell(row=row, column=16, value=datos_nuevos["cargo_linea"])  # P (16)
                ws_actas.cell(row=row, column=17, value=datos_nuevos["tipo"])         # Q (17)
                ws_actas.cell(row=row, column=26, value=datos_nuevos["tipo_cargo"])   # Z (26)
                
                contador_actualizaciones += 1
                print(f"✅ Fila {row} de Actas actualizada con éxito para el cargo: '{tipo_logia_acta}'")

    # Guardamos los cambios finales únicamente en el archivo de Actas
    if contador_actualizaciones > 0:
        wb_actas.save(RUTA_ACTAS_ENTREGA)
        print(f"💾 ¡Sincronización completada! Se arreglaron {contador_actualizaciones} registros en Actas de Entrega.")
    else:
        print("💤 No se encontraron actas pendientes que coincidieran con la Base Fabi actualizada.")
        wb_actas.close()

# Ejecutar el script
if __name__ == "__main__":
    sincronizar_actas_desde_base_fabi()