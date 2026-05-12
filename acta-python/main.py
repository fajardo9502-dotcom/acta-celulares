from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional
import base64
import os
from datetime import date
from openpyxl import load_workbook, Workbook
import threading

app = FastAPI(title="Acta Celulares API")

# --- CORS ---
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["POST", "OPTIONS"],
    allow_headers=["Content-Type"],
)

# --- CONFIGURACIÓN ---
EXCEL_PATH = r"C:\Users\1030650138\OneDrive - Colombiana de Comercio S.A\base actas de entrega.xlsx"
PDF_FOLDER = "PDFs"
os.makedirs(PDF_FOLDER, exist_ok=True)

# Lock para escritura thread-safe del Excel (igual que synchronized en Java)
excel_lock = threading.Lock()


# --- MODELO DE DATOS (equivalente al parsearJson de Java) ---
class DatosActa(BaseModel):
    Telefono: Optional[str] = ""
    IMEI1: Optional[str] = ""
    IMEI2: Optional[str] = ""
    MODELO: Optional[str] = ""
    C_Costos: Optional[str] = ""        # "C. Costos" → C_Costos (FastAPI/Pydantic)
    Supervisor: Optional[str] = ""
    Zona_o_Cargo: Optional[str] = ""    # "Zona o Cargo" → Zona_o_Cargo
    Codigo: Optional[str] = ""          # "Código"
    Cedula: Optional[str] = ""
    Funcionario: Optional[str] = ""
    Bateria: Optional[str] = "No"
    Cargador: Optional[str] = ""
    TipoEquipo: Optional[str] = ""
    Novedades: Optional[str] = ""
    Tipo_Plan: Optional[str] = ""       # "Tipo Plan"
    Costo_Plan: Optional[str] = ""      # "Costo Plan"
    Cuenta: Optional[str] = ""
    Nombre_Cuenta: Optional[str] = ""   # "Nombre Cuenta"
    Tipo_Cargo: Optional[str] = ""      # "Tipo Cargo"
    Tipo_Logia: Optional[str] = ""      # "Tipo Logia"
    pdfBase64: Optional[str] = ""
    firma_digital: Optional[str] = ""


def generar_nombre_pdf(datos: DatosActa) -> str:
    """Equivalente a generarNombrePdf() en Java."""
    cedula = "".join(c for c in (datos.Cedula or "sin_cedula") if c.isalnum())
    fecha = date.today().strftime("%Y%m%d")
    return f"acta_{cedula}_{fecha}.pdf"


def agregar_fila_excel(datos: DatosActa, ruta_pdf: str):
    """
    Equivalente a agregarFilaExcel() en Java.
    Columnas: A=No, B=Telefono, C=IMEI1, D=IMEI2, E=MODELO, F=C.Costos,
              G=Fecha Asignacion, H=UN2(vacío), I=UN, J=Supervisor,
              K=Zona/Cargo, L=Código, M=Cedula, N=Funcionario, O=Bateria,
              P=Cargador, Q=TipoEquipo, R=Estado, S=Novedades, T=Estado2,
              U=vacío, V=Tipo Plan, W=Costo Plan, X=Cuenta, Y=Nombre Cuenta,
              Z=Tipo Cargo, AA=Tipo Logia, AB=vacío, AC=Ruta PDF
    """
    with excel_lock:
        if os.path.exists(EXCEL_PATH):
            wb = load_workbook(EXCEL_PATH)
        else:
            wb = Workbook()
            wb.active.title = "base"

        ws = wb.active
        last_row = ws.max_row  # equivalente a sheet.getLastRowNum()
        numero_acta = last_row + 1
        fecha_hoy = date.today().strftime("%d/%m/%Y")

        nueva_fila = [
            numero_acta,                    # A - No (autoincremento)
            datos.Telefono,                 # B
            datos.IMEI1,                    # C
            datos.IMEI2,                    # D
            datos.MODELO,                   # E
            datos.C_Costos,                 # F
            fecha_hoy,                      # G - Fecha Asignacion
            "",                             # H - UN2 (vacío)
            datos.Zona_o_Cargo,             # I - UN
            datos.Supervisor,               # J
            datos.Zona_o_Cargo,             # K
            datos.Codigo,                   # L
            datos.Cedula,                   # M
            datos.Funcionario,              # N
            datos.Bateria,                  # O
            datos.Cargador,                 # P
            datos.TipoEquipo,               # Q
            "Activo",                       # R
            datos.Novedades,                # S
            "Activo",                       # T
            "",                             # U (vacío)
            datos.Tipo_Plan,                # V
            datos.Costo_Plan,               # W
            datos.Cuenta,                   # X
            datos.Nombre_Cuenta,            # Y
            datos.Tipo_Cargo,               # Z
            datos.Tipo_Logia,               # AA
            "",                             # AB (vacío)
            ruta_pdf,                       # AC
        ]

        ws.append(nueva_fila)
        wb.save(EXCEL_PATH)
        wb.close()


# --- ENDPOINT PRINCIPAL (equivalente a RecibirActa en Java) ---
@app.post("/api/acta")
async def recibir_acta(datos: DatosActa):
    try:
        # 1. Guardar PDF
        nombre_pdf = generar_nombre_pdf(datos)
        ruta_pdf = ""

        if datos.pdfBase64:
            pdf_b64 = datos.pdfBase64
            if "," in pdf_b64:
                pdf_b64 = pdf_b64.split(",", 1)[1]
            pdf_bytes = base64.b64decode(pdf_b64)
            ruta_pdf = os.path.join(PDF_FOLDER, nombre_pdf)
            with open(ruta_pdf, "wb") as f:
                f.write(pdf_bytes)
            print(f"  PDF guardado: {ruta_pdf}")

        # 2. Agregar fila al Excel
        agregar_fila_excel(datos, ruta_pdf)
        print(f"  Excel actualizado: {EXCEL_PATH}")

        return {"mensaje": "OK: PDF guardado y Excel actualizado"}

    except Exception as e:
        print(f"Error procesando acta: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# --- ARCHIVOS ESTÁTICOS (equivalente a ArchivoEstatico en Java) ---
# Sirve la carpeta actual como archivos estáticos (HTML, CSS, JS, imágenes)
app.mount("/", StaticFiles(directory=".", html=True), name="static")


# --- ARRANQUE DEL SERVIDOR ---
if __name__ == "__main__":
    import uvicorn
    print("----------------------------------------------")
    print("SERVIDOR LISTO EN: http://localhost:8080")
    print("----------------------------------------------")
    uvicorn.run(app, host="0.0.0.0", port=8080)
