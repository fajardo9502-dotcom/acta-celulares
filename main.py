# Librerías estándar de Python
import base64
import os
import threading
from datetime import date
from typing import Optional

# Librerías externas
import pandas as pd  # Usado para el cruce automático de datos corporativos
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel

# FastAPI
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles

app = FastAPI(title="Acta Celulares API")

# --- CORS ---
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["POST", "PUT", "OPTIONS"],
    allow_headers=["Content-Type"],
)

# --- CONFIGURACIÓN ---
EXCEL_PATH = r"C:\Users\1030650138\OneDrive - Colombiana de Comercio S.A\base actas de entrega.xlsx"
RUTA_BASE_ALIMENTADORA = r"C:\Users\1030650138\OneDrive - Colombiana de Comercio S.A\report_Planta_de_Personal_Completa__activos_e_inactivos_--colombiana--1010236537--_93bb93e0-54b6-46de-9477-06f6b4e6b9c3 (1).xlsx"
RUTA_BASE_TIPOLOGIAS = r"C:\Users\1030650138\OneDrive - Colombiana de Comercio S.A\Base Fabi.xlsx"
RUTA_BASE_EQUIPOS_HISTORICO = r"C:\Users\1030650138\OneDrive - Colombiana de Comercio S.A\Base Historicos Fabi.xlsx"
PDF_FOLDER = "PDFs"
os.makedirs(PDF_FOLDER, exist_ok=True)

# Lock para escritura thread-safe del Excel
excel_lock = threading.RLock()


# ================================================================
# UTILIDAD: Limpieza estricta de cédula
# ================================================================
def limpiar_cedula(valor) -> str:
    """
    Convierte cualquier valor de cédula a texto numérico estricto.
    Elimina espacios, el '.0' que mete Excel al leer números,
    y la notación científica que puede generar Pandas.
    Ejemplos:
        "1030650138.0"  -> "1030650138"
        " 1030650138 "  -> "1030650138"
        1030650138.0    -> "1030650138"
    """
    if valor is None:
        return ""
    texto = str(valor).strip()
    try:
        # Esto convierte "1030650138.0" o 1.03e+09 a entero limpio
        texto = str(int(float(texto)))
    except (ValueError, OverflowError):
        pass
    # Quitar puntos de miles y espacios residuales
    texto = texto.replace(".", "").replace(",", "").replace(" ", "")
    return texto


# ================================================================
# UTILIDADES DE TEXTO
# ================================================================
def limpiar_mayuscula(texto) -> str:
    if texto:
        return str(texto).strip().upper()
    return ""

def limpiar_titulo(texto) -> str:
    if texto:
        return str(texto).strip().title()
    return ""


# ================================================================
# MODELOS DE DATOS
# ================================================================
class DatosActa(BaseModel):
    Telefono: Optional[str] = ""
    IMEI1: Optional[str] = ""
    IMEI2: Optional[str] = ""
    MODELO: Optional[str] = ""
    marca: Optional[str] = ""
    C_Costos: Optional[str] = ""
    UN: Optional[str] = ""
    Supervisor: Optional[str] = ""
    Zona_o_Cargo: Optional[str] = ""
    Codigo: Optional[str] = ""
    Cedula: Optional[str] = ""
    Funcionario: Optional[str] = ""
    Bateria: Optional[str] = "No"
    Cargador: Optional[str] = ""
    TIPOEQUIPO: Optional[str] = ""
    Novedades: Optional[str] = ""
    Tipo_Plan: Optional[str] = ""
    Costo_Plan: Optional[str] = ""
    Cuenta: Optional[str] = ""
    Nombre_Cuenta: Optional[str] = ""
    Tipo_Cargo: Optional[str] = ""
    Tipo_Logia: Optional[str] = ""
    pdfBase64: Optional[str] = ""
    firma_digital: Optional[str] = ""


class DatosDescuento(BaseModel):
    Cedula: Optional[str] = ""
    Funcionario: Optional[str] = ""
    IMEI1: Optional[str] = ""
    MODELO: Optional[str] = ""
    ValorDescuento: Optional[str] = ""
    Novedades: Optional[str] = ""


# ================================================================
# UTILIDAD: Búsqueda flexible de columnas
# ================================================================
def buscar_columna(df_columns, nombre_base: str) -> Optional[str]:
    """
    Busca una columna por nombre exacto primero. Si no la encuentra,
    busca cualquier columna que EMPIECE con el nombre base (útil para
    reportes tipo SuccessFactors que numeran columnas como
    'Nombre de usuario (0)', 'Código de cargo Nombre de cargo (1)', etc.
    y donde el número puede cambiar entre exportaciones).
    """
    if nombre_base in df_columns:
        return nombre_base
    for col in df_columns:
        if col.strip().startswith(nombre_base):
            return col
    return None


# ================================================================
# FUNCIÓN DE CRUCE POR CÉDULA (Base Alimentadora)
# ================================================================
def procesar_cruce_por_cedula(cedula_usuario: str) -> dict:
    """
    Busca al empleado en la base alimentadora por su cédula.
    Retorna un diccionario con cuatro claves:
        - "UN2":            parte corta del centro de costos (ej: "104")
        - "C_COSTOS_LARGO": parte larga del centro de costos (ej: "944058")
        - "UN":             nombre de la unidad organizativa
        - "Tipo_Logia":     Nombre del cargo extraído de la columna H (Base Alimentadora)

    IMPORTANTE: Si NO encuentra al empleado, las claves vienen
    como None para que la función de escritura active la salvaguarda de datos.
    """
    # Agregamos "Tipo_Logia" inicializado en None para la salvaguarda
    resultado = {"UN2": None, "C_COSTOS_LARGO": None, "UN": None, "Tipo_Logia": None}

    cedula_limpia = limpiar_cedula(cedula_usuario)
    if not cedula_limpia:
        print("   Cedula vacia, se omite el cruce.")
        return resultado

    if not os.path.exists(RUTA_BASE_ALIMENTADORA):
        print(f"   Base alimentadora no encontrada en: {RUTA_BASE_ALIMENTADORA}")
        return resultado

    try:
        df = pd.read_excel(RUTA_BASE_ALIMENTADORA, sheet_name="Planta de Personal_Completa (ac", dtype=str)
        df.columns = df.columns.str.strip()

        columna_cedula = 'Nombre de usuario (0)'

        if columna_cedula not in df.columns:
            print(f"  No se encontro columna de cedula. Columnas disponibles: {list(df.columns)}")
            return resultado

        df["_cedula_limpia"] = df[columna_cedula].apply(limpiar_cedula)
        coincidencia = df[df["_cedula_limpia"] == cedula_limpia]

        if coincidencia.empty:
            print(f"   Cedula '{cedula_limpia}' NO encontrada en la base alimentadora.")
            return resultado  

        fila = coincidencia.iloc[0]

        # --- [CRUCE ORIGINAL] Centro de Costos ---
        columna_cc = buscar_columna(df.columns, "Centro de costos Código")
        if columna_cc:
            valor_cc = str(fila.get(columna_cc, "")).strip()
            if valor_cc.lower() in ("nan", "none", ""):
                valor_cc = ""

            if "-" in valor_cc:
                un2_parte, largo_parte = valor_cc.split("-", 1)
                resultado["UN2"] = un2_parte.strip()
                resultado["C_COSTOS_LARGO"] = largo_parte.strip()
            elif valor_cc:
                resultado["C_COSTOS_LARGO"] = valor_cc
                resultado["UN2"] = ""
        else:
            print(f"   Columna 'Centro de costos Código' no encontrada en la base alimentadora. "
                  f"Columnas disponibles: {list(df.columns)}")

        # --- [CRUCE ORIGINAL] Nombre de la Unidad (UN) ---
        columna_un = buscar_columna(df.columns, "Unidad Estratégica de Negocio Nombre")
        if columna_un:
            valor_un = str(fila.get(columna_un, "")).strip()
            resultado["UN"] = "" if valor_un.lower() in ("nan", "none") else valor_un.upper()
        else:
            resultado["UN"] = ""

        # --- [NUEVO CRUCE] Extraer columna H (Cargo) ---
        columna_cargo = buscar_columna(df.columns, "Código de cargo Nombre de cargo")
        if columna_cargo:
            valor_cargo = str(fila.get(columna_cargo, "")).strip()
            resultado["Tipo_Logia"] = "" if valor_cargo.lower() in ("nan", "none") else valor_cargo.upper()
        else:
            print(f"   Columna 'Código de cargo Nombre de cargo' no encontrada en la base alimentadora. "
                  f"Columnas disponibles: {list(df.columns)}. Se asigna vacio.")
            resultado["Tipo_Logia"] = ""

        print(f"   [Cruce OK] Cedula: {cedula_limpia} -> "
              f"UN2: {resultado['UN2']} | C.Costos: {resultado['C_COSTOS_LARGO']} | Cargo (AA): {resultado['Tipo_Logia']}")
        return resultado

    except Exception as e:
        print(f"   Error en procesar_cruce_por_cedula: {e}")
        return resultado

# ================================================================
# FUNCIÓN DE CRUCE PARA BASE FABI (Base Alimentadora)
# ================================================================
def obtener_datos_tipologia_fabi(tipo_logia_buscado: str) -> dict:
    """
    Segundo Cruce: Busca el Tipo Logia en la 'Hoja1' de la 'Base Fabi' y extrae:
    - Tipo Cargo Final  -> Columna B
    - Tipo Final        -> Columna C
    - Cargo Linea Final -> Columna D
    """
    resultado = {"Tipo_Cargo_Final": "", "Tipo_Final": "", "Cargo_Linea_Final": "", "Falta_En_Base": True}
    
    if not tipo_logia_buscado:
        return resultado
        
    try:
        if not os.path.exists(RUTA_BASE_TIPOLOGIAS):
            print(f"    Base Fabi no encontrada en: {RUTA_BASE_TIPOLOGIAS}")
            return resultado
            
        # 🎯 FUERZA A PANDAS A LEER ÚNICAMENTE LA PESTAÑA "Hoja1"
        df_fabi = pd.read_excel(RUTA_BASE_TIPOLOGIAS, sheet_name="Hoja1")
        
        # Limpieza de espacios fantasmas en los encabezados por si acaso
        df_fabi.columns = df_fabi.columns.astype(str).str.strip()
        
        # Nombres exactos de las columnas en tu "Hoja1"
        col_llave = "Tipo Logia Final"
        col_cargo = "Tipo Cargo Final"
        col_tipo = "Tipo Final"
        col_linea = "Cargo Linea Final"
        
        if col_llave not in df_fabi.columns:
            print(f"    Error: La columna '{col_llave}' no existe en Hoja1. Columnas: {list(df_fabi.columns)}")
            return resultado
            
        # Homogeneizar cadenas para evitar fallos por minúsculas o espacios libres
        df_fabi[col_llave] = df_fabi[col_llave].astype(str).str.strip().str.upper()
        busqueda = str(tipo_logia_buscado).strip().upper()
        
        fila_coincidente = df_fabi[df_fabi[col_llave] == busqueda]
        
        if not fila_coincidente.empty:
            fila = fila_coincidente.iloc[0]
            
            val_tipo_cargo = str(fila.get(col_cargo, "")).strip()
            val_tipo_final = str(fila.get(col_tipo, "")).strip()
            val_cargo_linea = str(fila.get(col_linea, "")).strip()
            
            # Guardamos los resultados limpios en mayúsculas
            resultado["Tipo_Cargo_Final"] = "" if val_tipo_cargo.lower() in ("nan", "none") else val_tipo_cargo.upper()
            resultado["Tipo_Final"] = "" if val_tipo_final.lower() in ("nan", "none") else val_tipo_final.upper()
            resultado["Cargo_Linea_Final"] = "" if val_cargo_linea.lower() in ("nan", "none") else val_cargo_linea.upper()
            resultado["Falta_En_Base"] = False  
            
            print(f"    [Cruce Base Fabi OK] '{busqueda}' encontrado en Hoja1.")
            return resultado
        else:
            print(f"    ALERTA: El cargo '{busqueda}' NO existe en la columna '{col_llave}' de la Hoja1.")
            return resultado
            
    except Exception as e:
        print(f"    Error crítico en cruce con Base Fabi (Hoja1): {str(e)}")
        return resultado
    
# ================================================================
# GENERADOR DE NOMBRE DE PDF
# ================================================================
def generar_nombre_pdf(datos: DatosActa) -> str:
    cedula = "".join(c for c in (datos.Cedula or "sin_cedula") if c.isalnum())
    fecha = date.today().strftime("%Y%m%d")
    return f"acta_{cedula}_{fecha}.pdf"

# ================================================================
# FUNCIÓN DE CRUCE POR IMEI (Buscando en la Base de Equipos fabi)
# VERSIONES CORREGIDAS CON OPENPYXL (Sincronización Total)
# ================================================================

def obtener_anterior_usuario_por_imei(imei_buscado: str) -> str:
    """
    Busca un IMEI en la Base de Históricos y retorna los últimos 2 usuarios
    registrados en formato: "USUARIO_VIEJO <--- USUARIO_RECIENTE"
    """
    if not imei_buscado or not os.path.exists(RUTA_BASE_EQUIPOS_HISTORICO):
        return ""

    try:
        # Usamos el candado global para evitar bloqueos
        with excel_lock:
            wb = load_workbook(RUTA_BASE_EQUIPOS_HISTORICO, read_only=True)
            try:
                ws = wb.active
                
                # Obtenemos las cabeceras de la fila 1
                cabeceras = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
                
                def buscar_indice(nombres_posibles):
                    for nombre in nombres_posibles:
                        for i, cabecera in enumerate(cabeceras):
                            if cabecera.lower() == nombre.lower():
                                return i + 1
                    return None

                # Mapeo dinámico de columnas por nombre
                col_imei_idx = buscar_indice(["imei1", "imei"])
                col_cedula_idx = buscar_indice(["cedula", "cédula", "documento"])
                col_nombre_idx = buscar_indice(["nombre", "funcionario", "nombre completo"])
                col_fecha_idx = buscar_indice(["fecha", "fecha_entrega"])

                if not col_imei_idx:
                    print(f"   [Cruce IMEI Error] No se encontró columna IMEI en {cabeceras}")
                    return ""

                busqueda = str(imei_buscado).strip()
                coincidencias = []

                # Recorremos todas las filas buscando el IMEI
                for row in range(2, ws.max_row + 1):
                    val_imei = str(ws.cell(row=row, column=col_imei_idx).value or "").strip()
                    
                    if val_imei == busqueda:
                        cedula = str(ws.cell(row=row, column=col_cedula_idx).value or "").strip() if col_cedula_idx else ""
                        nombre = str(ws.cell(row=row, column=col_nombre_idx).value or "").strip() if col_nombre_idx else ""
                        fecha = str(ws.cell(row=row, column=col_fecha_idx).value or "").strip() if col_fecha_idx else ""

                        # Limpieza de valores nulos o "nan"
                        cedula = "" if cedula.lower() in ("nan", "none") else cedula
                        nombre = "" if nombre.lower() in ("nan", "none") else nombre.upper()
                        fecha = "" if fecha.lower() in ("nan", "none") else fecha

                        if cedula or nombre:
                            coincidencias.append(f"{cedula} | {nombre} | {fecha}")

                # Si encontramos registros previos
                if coincidencias:
                    # '[-2:]' extrae exactamente los 2 últimos elementos del arreglo
                    ultimas_dos = coincidencias[-2:]
                    
                    if len(ultimas_dos) == 2:
                        historial_formateado = f"{ultimas_dos[0]}  <---  {ultimas_dos[1]}"
                    else:
                        historial_formateado = ultimas_dos[0]
                    
                    print(f"   [Histórico OK] Registros encontrados: {historial_formateado}")
                    return historial_formateado
                
                return ""
            finally:
                wb.close()
                
    except Exception as e:
        print(f"   Error leyendo histórico para IMEI {imei_buscado}: {str(e)}")
        return ""


# ================================================================
# ACTUALIZACIÓN DEL EXCEL DE ACTAS (hoja principal "base")
# ================================================================

def actualizar_fila_excel_actas(datos: DatosActa, ruta_pdf: str):

    """
    Actualiza o crea una fila en el Excel de actas basada en la cédula del empleado.
    
    GARANTÍAS:
    1. Novedades SIEMPRE va a columna 19 (S)
    2. Tipo_Logia va a columna 27 (AA)
    3. Cruce por Cargo (Base Fabi) se inyecta en las columnas P, Q y Z según lo solicitado.
    4. Salvaguarda activada: Si cédula no existe en planta, F y H NO se tocan.
    5. Anterior Usuario se inyecta en columna 28 (AB) obtenido cruzando por IMEI.
    """
    with excel_lock:
        if os.path.exists(EXCEL_PATH):
            wb = load_workbook(EXCEL_PATH)
        else:
            wb = Workbook()
            wb.active.title = "base"
 
        ws = wb.active
        cedula_a_buscar = limpiar_cedula(datos.Cedula)
        empleado_encontrado = False
 
        # === PASO 1: CRUCE CON BASE ALIMENTADORA (Por cédula) ===
        info_unidades = procesar_cruce_por_cedula(cedula_a_buscar)
        cruce_exitoso = info_unidades["UN2"] is not None
 
        # === PASO 2: DETERMINAR TIPO_LOGIA FINAL ===
        tipo_logia_final = ""
        if cruce_exitoso and info_unidades.get("Tipo_Logia"):
            tipo_logia_final = info_unidades["Tipo_Logia"]
            print(f"   [Tipo_Logia] Usando valor de base alimentadora: {tipo_logia_final}")
        elif datos.Tipo_Logia:
            tipo_logia_final = datos.Tipo_Logia
            print(f"   [Tipo_Logia] Usando valor del formulario web: {tipo_logia_final}")
        else:
            print(f"   [Tipo_Logia] Sin valor disponible (quedará vacío)")

        # === PASO 2.5: EJECUTAR EL SEGUNDO CRUCE (BASE FABI) ===
        datos_fabi = obtener_datos_tipologia_fabi(tipo_logia_final)
        falta_cargo_en_base = datos_fabi["Falta_En_Base"]
 
        # 🎯 === PASO 2.8: EJECUTAR EL TERCER CRUCE (ANTERIOR USUARIO POR IMEI) ===
        anterior_usuario_data = obtener_anterior_usuario_por_imei(datos.IMEI1)

        # === PASO 3: BUSCAR FILA EXISTENTE POR CÉDULA ===
        target_row = None
        if ws.max_row >= 2 and cedula_a_buscar:
            for row in range(2, ws.max_row + 1):
                cedula_celda = limpiar_cedula(ws.cell(row=row, column=13).value)
 
                if cedula_celda == cedula_a_buscar:
                    
                    # FILA ENCONTRADA: ACTUALIZAR EXISTENTE
                    fecha_hoy = date.today().strftime("%d/%m/%Y")
 
                    # Datos del formulario web (SIEMPRE se actualizan)
                    ws.cell(row=row, column=2,  value=limpiar_mayuscula(datos.Telefono))         # B
                    ws.cell(row=row, column=3,  value=limpiar_mayuscula(datos.IMEI1))            # C
                    ws.cell(row=row, column=4,  value=limpiar_mayuscula(datos.IMEI2))            # D
                    ws.cell(row=row, column=5,  value=limpiar_mayuscula(f"{datos.marca} - {datos.MODELO}")) # E
                    ws.cell(row=row, column=7,  value=fecha_hoy)                                 # G
                    ws.cell(row=row, column=9,  value=limpiar_mayuscula(datos.UN))                                   # I
                    ws.cell(row=row, column=10, value=limpiar_titulo(datos.Supervisor))          # J
                    ws.cell(row=row, column=11, value=limpiar_mayuscula(datos.Zona_o_Cargo))     # K
                    ws.cell(row=row, column=12, value=limpiar_mayuscula(datos.Codigo))           # L
                    ws.cell(row=row, column=14, value=limpiar_titulo(datos.Funcionario))         # N
                    ws.cell(row=row, column=15, value=limpiar_mayuscula(datos.Bateria))          # O
                    
                    # COLUMNA P (16): Cargo Linea Final de Base Fabi
                    ws.cell(row=row, column=16, value=datos_fabi["Cargo_Linea_Final"])           # P
                    
                    # COLUMNA Q (17): Tipo Final de Base Fabi
                    ws.cell(row=row, column=17, value=datos_fabi["Tipo_Final"])                  # Q
                    
                    ws.cell(row=row, column=18, value=limpiar_mayuscula(datos.TIPOEQUIPO))       # R
                    
                    # NOVEDADES EN COLUMNA 19 (S) - GARANTIZADO
                    ws.cell(row=row, column=19, value=limpiar_mayuscula(datos.Novedades))        # S
                    print(f"   [Columna S] Novedades guardadas: {datos.Novedades}")
                    
                    ws.cell(row=row, column=22, value=limpiar_mayuscula(datos.Tipo_Plan))        # V
                    ws.cell(row=row, column=23, value=limpiar_mayuscula(datos.Costo_Plan))       # W
                    ws.cell(row=row, column=24, value=limpiar_mayuscula(datos.Cuenta))           # X
                    ws.cell(row=row, column=25, value=limpiar_mayuscula(datos.Nombre_Cuenta))    # Y
                    
                    # COLUMNA Z (26): Tipo Cargo Final de Base Fabi
                    ws.cell(row=row, column=26, value=datos_fabi["Tipo_Cargo_Final"])            # Z
                    
                    # TIPO_LOGIA EN COLUMNA 27 (AA) - GARANTIZADO
                    ws.cell(row=row, column=27, value=limpiar_mayuscula(tipo_logia_final))       # AA
                    print(f"   [Columna AA] Tipo_Logia guardado: {tipo_logia_final}")
                    
                    #  COLUMNA 28 (AB): Anterior Usuario obtenido por el cruce por IMEI
                    ws.cell(row=row, column=28, value=anterior_usuario_data)                     # AB
                    print(f"   [Columna AB] Anterior Usuario guardado: {anterior_usuario_data}")

                    ws.cell(row=row, column=29, value=ruta_pdf)                                  # AC
 
                    # SALVAGUARDA: Solo tocar F y H si cruce fue exitoso
                    if cruce_exitoso:
                        ws.cell(row=row, column=6, value=info_unidades["C_COSTOS_LARGO"])  # F
                        ws.cell(row=row, column=8, value=info_unidades["UN2"])              # H
                        print(f"   [Cruce Exitoso] Columnas F, H actualizadas.")
                    else:
                        print(f"    [Salvaguarda] Cédula no en planta. Columnas F, H preservadas.")
 
                    print(f"   Fila {row} ACTUALIZADA para cédula: {cedula_a_buscar}")
                    empleado_encontrado = True
                    target_row = row
                    break
 
        # === PASO 4: CREAR FILA NUEVA SI NO EXISTE ===
        if not empleado_encontrado:
            numero_acta = ws.max_row + 1
            target_row = ws.max_row + 1
            ws.cell(row=target_row, column=1, value=target_row)  # ID automático en columna A
            print(f"   Fila NUEVA ({target_row}) asignada para cédula: {cedula_a_buscar}")
            fecha_hoy = date.today().strftime("%d/%m/%Y")
 
            # Armar la lista nueva_fila con TODOS los datos en orden exacto
            nueva_fila = [
                numero_acta,                                                     # A  (1)
                limpiar_mayuscula(datos.Telefono),                               # B  (2)
                limpiar_mayuscula(datos.IMEI1),                                  # C  (3)
                limpiar_mayuscula(datos.IMEI2),                                  # D  (4)
                limpiar_mayuscula(f"{datos.marca} - {datos.MODELO}"),            # E  (5)
                info_unidades["C_COSTOS_LARGO"] if cruce_exitoso else "",        # F  (6)
                fecha_hoy,                                                       # G  (7)
                info_unidades["UN2"] if cruce_exitoso else "",                   # H  (8)
                limpiar_mayuscula(datos.UN),                                     # I  (9)
                limpiar_titulo(datos.Supervisor),                                # J  (10)
                limpiar_mayuscula(datos.Zona_o_Cargo),                           # K  (11)
                limpiar_mayuscula(datos.Codigo),                                 # L  (12)
                cedula_a_buscar,                                                 # M  (13)
                limpiar_titulo(datos.Funcionario),                               # N  (14)
                limpiar_mayuscula(datos.Bateria),                                # O  (15)
                datos_fabi["Cargo_Linea_Final"],                                 # P  (16)
                datos_fabi["Tipo_Final"],                                        # Q  (17)
                limpiar_mayuscula(datos.TIPOEQUIPO),                             # R  (18)
                limpiar_mayuscula(datos.Novedades),                              # S  (19)
                "",                                                              # T  (20)
                "",                                                              # U  (21)
                limpiar_mayuscula(datos.Tipo_Plan),                              # V  (22)
                limpiar_mayuscula(datos.Costo_Plan),                             # W  (23)
                limpiar_mayuscula(datos.Cuenta),                                 # X  (24)
                limpiar_mayuscula(datos.Nombre_Cuenta),                          # Y  (25)
                datos_fabi["Tipo_Cargo_Final"],                                  # Z  (26)
                limpiar_mayuscula(tipo_logia_final),                             # AA (27)
                anterior_usuario_data,                                           # AB (28)  ← 🎯 NUEVO: Anterior Usuario por IMEI
                ruta_pdf,                                                        # AC (29)
            ]
            
            ws.append(nueva_fila)
            print(f"    Fila NUEVA creada para cédula: {cedula_a_buscar}")
 
        # === PASO 5: GUARDAR ===
        wb.save(EXCEL_PATH)
        wb.close()
        print(f"    Excel guardado exitosamente")
 
        #  === PASO 5.5: REGISTRAR ASIGNACIÓN ACTUAL EN LA BASE DE HISTÓRICOS ===
        # Esto guarda al nuevo dueño para que sirva de "Anterior Usuario" en el futuro
        registrar_en_base_historicos(
            imei=datos.IMEI1,
            cedula=cedula_a_buscar,
            nombre=datos.Funcionario
        )

        # === PASO 6: RETORNAR SI EL CARGO FALTA EN LA BASE FABI ===
        return falta_cargo_en_base
    
# ================================================================
# FUNCIÓN PARA GUARDAR/REGISTRAR EN LA BASE DE HISTÓRICOS
# ================================================================
def registrar_en_base_historicos(imei: str, cedula: str, nombre: str):
    """
    Agrega un nuevo registro al final del archivo de históricos
    con el IMEI, Cédula, Nombre del nuevo usuario y la Fecha de hoy.
    """
    if not imei:
        print("    [Histórico] No se puede registrar: IMEI vacío.")
        return

    if not os.path.exists(RUTA_BASE_EQUIPOS_HISTORICO):
        print(f"    [Histórico] Error: No se encontró el archivo de históricos para actualizar: {RUTA_BASE_EQUIPOS_HISTORICO}")
        return

    try:
        # Aseguramos bloquear el proceso para que no se pise con otros hilos (usando openpyxl con tu lock)
        with excel_lock:
            wb = load_workbook(RUTA_BASE_EQUIPOS_HISTORICO)
            # Abrimos la primera hoja por defecto (index 0)
            ws = wb.active 
            
            # Formateamos la fecha de hoy para el registro
            fecha_hoy = date.today().strftime("%d/%m/%Y")
            
            # Limpiamos los datos para guardarlos impecables
            imei_limpio = str(imei).strip()
            cedula_limpia = str(cedula).strip()
            nombre_limpio = str(nombre).strip().upper() # Siempre en mayúsculas
            
            # Buscamos los índices de las columnas por sus nombres en la primera fila (Cabecera)
            cabeceras = [str(cell.value).strip() for cell in ws[1]]
            
            col_imei_idx = cabeceras.index("IMEI1") + 1 if "IMEI1" in cabeceras else None
            col_cedula_idx = cabeceras.index("Cedula") + 1 if "Cedula" in cabeceras else None
            col_nombre_idx = cabeceras.index("Nombre") + 1 if "Nombre" in cabeceras else None
            col_fecha_idx = cabeceras.index("Fecha") + 1 if "Fecha" in cabeceras else None

            # Si el archivo de históricos no tiene estas columnas exactas, las creamos o usamos por defecto
            if not all([col_imei_idx, col_cedula_idx, col_nombre_idx, col_fecha_idx]):
                print("    [Histórico] Advertencia: Las columnas estándar no coinciden. Guardando en orden A, B, C, D.")
                col_imei_idx, col_cedula_idx, col_nombre_idx, col_fecha_idx = 1, 2, 3, 4
                
            # Determinamos la siguiente fila disponible
            nueva_fila = ws.max_row + 1
            
            # Escribimos los datos del nuevo asignado en el histórico
            ws.cell(row=nueva_fila, column=col_imei_idx, value=imei_limpio)
            ws.cell(row=nueva_fila, column=col_cedula_idx, value=cedula_limpia)
            ws.cell(row=nueva_fila, column=col_nombre_idx, value=nombre_limpio)
            ws.cell(row=nueva_fila, column=col_fecha_idx, value=fecha_hoy)
            
            # Guardamos el archivo Excel de históricos
            wb.save(RUTA_BASE_EQUIPOS_HISTORICO)
            wb.close()
            print(f"    [Histórico OK] Se registró nueva asignación: {imei_limpio} -> {cedula_limpia} | {nombre_limpio} | {fecha_hoy}")

    except Exception as e:
        print(f"    [Histórico Error] No se pudo guardar el registro de histórico: {str(e)}")


# ================================================================
# REGISTRO DE DESCUENTOS (hoja "Descuentos")
# ================================================================
def agregar_fila_descuento(datos: DatosDescuento):
    with excel_lock:
        if os.path.exists(EXCEL_PATH):
            wb = load_workbook(EXCEL_PATH)
        else:
            wb = Workbook()
            wb.active.title = "base"

        if "Descuentos" in wb.sheetnames:
            ws = wb["Descuentos"]
        else:
            ws = wb.create_sheet(title="Descuentos")
            ws.append(["Fecha", "Cedula", "Funcionario", "IMEI", "Modelo", "Valor Descuento", "Motivo"])

        fecha_hoy = date.today().strftime("%d/%m/%Y")
        ws.append([
            fecha_hoy,
            limpiar_cedula(datos.Cedula),
            limpiar_titulo(datos.Funcionario),
            datos.IMEI1,
            datos.MODELO,
            datos.ValorDescuento,
            datos.Novedades,
        ])
        wb.save(EXCEL_PATH)
        wb.close()


# ================================================================
# ENDPOINTS API
# ================================================================
@app.put("/api/acta")
async def recibir_acta(datos: DatosActa):
    try:
        nombre_pdf = generar_nombre_pdf(datos)
        ruta_pdf = ""

        if datos.pdfBase64:
            pdf_b64 = datos.pdfBase64
            if "," in pdf_b64:
                pdf_b64 = pdf_b64.split(",", 1)[1]
            pdf_bytes = base64.b64decode(pdf_b64)
            ruta_pdf = os.path.join(PDF_FOLDER, nombre_pdf)
            with open(ruta_pdf, "wb") as f:
                f.write(pdf_bytes)
            print(f"   PDF guardado: {ruta_pdf}")

        # Guardamos el booleano que nos dice si el cargo no existía
        cargo_no_existe = actualizar_fila_excel_actas(datos, ruta_pdf)
        print(f"   Excel actualizado: {EXCEL_PATH}")
        
        # --- RESPUESTA DINÁMICA ---
        if cargo_no_existe:
            cargo_afectado = (datos.Tipo_Logia or "DESCONOCIDO").upper()
            return {
                "status": "PROCESADO_CON_ADVERTENCIA",
                "mensaje": f"OK: Registro actualizado, pero el cargo '{cargo_afectado}' NO se encuentra en la Base Fabi y debe ser agregado.",
                "alerta_base_fabi": True,
                "cargo_faltante": cargo_afectado
            }
        
        return {
            "status": "OK",
            "mensaje": "OK: Registro actualizado y PDF guardado correctamente.",
            "alerta_base_fabi": False
        }

    except Exception as e:
        import traceback
        print(f"Error procesando acta: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/descuento")
async def recibir_descuento(datos: DatosDescuento):
    try:
        print(f"--> Peticion recibida en /api/descuento para: {datos.Funcionario}")
        agregar_fila_descuento(datos)
        return {"mensaje": "OK: Descuento registrado en el sistema"}
    except Exception as e:
        import traceback
        print(f"Error procesando el descuento: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ================================================================
# ARCHIVOS ESTÁTICOS
# ================================================================
app.mount("/", StaticFiles(directory="templates", html=True), name="static")


# ================================================================
# ARRANQUE
# ================================================================
if __name__ == "__main__":
    import uvicorn
    print("----------------------------------------------")
    print("SERVIDOR LISTO EN: http://localhost:8080")
    print("----------------------------------------------")
    uvicorn.run(app, host="0.0.0.0", port=8080)